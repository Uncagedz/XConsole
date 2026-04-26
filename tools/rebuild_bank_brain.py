from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import pymupdf as fitz
import requests
import trafilatura
from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from openpyxl import load_workbook

from routeone_bank_docs_sync import DEFAULT_BANK_ROOT, ROOT, build_bank_aliases, infer_bank, rel, sanitize_filename

DEFAULT_DECODED_DIR = ROOT / 'runtime' / 'routeone_docs' / 'decoded'
DEFAULT_INDEX_PATH = ROOT / 'runtime' / 'routeone_docs' / 'decoded_index.json'
DEFAULT_PROFILES_PATH = ROOT / 'data' / 'bank_profiles.generated.json'
DEFAULT_SALES_BANKS_PATH = ROOT / 'sales-assistant' / 'data' / 'banks.json'
DEFAULT_LINK_CACHE_DIR = ROOT / 'runtime' / 'routeone_docs' / 'linked_cache'

TEXT_EXTENSIONS = {'.pdf', '.txt', '.html', '.htm', '.csv', '.json', '.xlsx', '.xlsm', '.docx'}
RELEVANT_LINK_RE = re.compile(
    r'(pdf|rate|form|program|guideline|approval|credit|funding|stip|lender|matrix|bulletin|bankruptcy|repo)',
    re.IGNORECASE,
)
STIP_KEYWORDS = [
    'proof of income',
    'proof of residence',
    'stip',
    'driver license',
    'driver\'s license',
    'insurance',
    'bank statement',
    'pay stub',
    'utility bill',
    'trade payoff',
    'references',
    'social security',
]
RULE_KEYWORDS = ['bankruptcy', 'repo', 'repossession', 'charge-off', 'collection', 'first time buyer', 'seasoning']

TIER_DEFAULTS: dict[str, dict[str, Any]] = {
    'prime': {
        'min_score': 680,
        'max_ltv': 120,
        'max_pti': 15,
        'max_dti': 42,
        'max_derogatories': 2,
        'max_utilization': 65,
        'max_term_months': 84,
        'weight': 1.8,
    },
    'near_prime': {
        'min_score': 620,
        'max_ltv': 128,
        'max_pti': 17,
        'max_dti': 48,
        'max_derogatories': 3,
        'max_utilization': 82,
        'max_term_months': 84,
        'weight': 1.5,
    },
    'subprime': {
        'min_score': 540,
        'max_ltv': 140,
        'max_pti': 20,
        'max_dti': 55,
        'max_derogatories': 6,
        'max_utilization': 95,
        'max_term_months': 75,
        'weight': 1.2,
    },
    'credit_union': {
        'min_score': 660,
        'max_ltv': 115,
        'max_pti': 14,
        'max_dti': 40,
        'max_derogatories': 2,
        'max_utilization': 65,
        'max_term_months': 72,
        'weight': 1.7,
    },
    'captive': {
        'min_score': 620,
        'max_ltv': 130,
        'max_pti': 17,
        'max_dti': 48,
        'max_derogatories': 3,
        'max_utilization': 80,
        'max_term_months': 84,
        'weight': 1.6,
    },
}

BANK_TIER_HINTS = {
    'Ally': 'near_prime',
    'American_Credit_Acceptance_LLC': 'subprime',
    'AMERICREDIT': 'captive',
    'Axos_Bank': 'prime',
    'Bank_of_America': 'prime',
    'Cal_Automotive': 'near_prime',
    'CapitalOne': 'near_prime',
    'Chase': 'prime',
    'Dade_County_Federal_Credit_Union': 'credit_union',
    'Exeter_Finance': 'subprime',
    'FCA_Mastercard': 'captive',
    'Fifth_Third_Bank_National_Association': 'prime',
    'First_Help_Financial': 'subprime',
    'Foursight_Capital': 'subprime',
    'Global_Lending_Services': 'subprime',
    'GoFi_LLC': 'subprime',
    'GTE_Federal_Credit_Union': 'credit_union',
    'Mid_Florida_Credit_Union': 'credit_union',
    'PNC_Bank': 'prime',
    'Santander': 'near_prime',
    'Space_Coast_Credit_Union': 'credit_union',
    'Teachers_Federal_Credit_Union': 'credit_union',
    'Tropical_Financial_Credit_Union': 'credit_union',
    'US_Bank': 'prime',
    'Valley_National_Bank': 'prime',
    'Valley_Strong_Credit_Union': 'credit_union',
    'Wells_Fargo_Auto': 'prime',
    'Westlake_Financial_Services': 'subprime',
}

METRIC_RULES: dict[str, dict[str, Any]] = {
    'min_score': {
        'fallback': lambda bank, tier: TIER_DEFAULTS[tier]['min_score'],
        'bounds': (300, 850),
        'pick': min,
        'patterns': [
            r'(?:minimum|min\.?|score\s+floor|cutoff|fico\s+score|credit\s+score)[^\d]{0,18}([3-8]\d{2})',
        ],
    },
    'max_ltv': {
        'fallback': lambda bank, tier: TIER_DEFAULTS[tier]['max_ltv'],
        'bounds': (80, 200),
        'pick': max,
        'patterns': [
            r'(?:max(?:imum)?\s*)?(?:ltv|loan\s*to\s*value|advance(?:\s+amount)?)[^\d]{0,18}([8-9]\d|1\d\d|200)(?:\.\d+)?\s*%',
        ],
    },
    'max_pti': {
        'fallback': lambda bank, tier: TIER_DEFAULTS[tier]['max_pti'],
        'bounds': (5, 30),
        'pick': max,
        'patterns': [
            r'(?:max(?:imum)?\s*)?(?:pti|payment\s*to\s*income)[^\d]{0,18}(\d{1,2}(?:\.\d+)?)\s*%',
        ],
    },
    'max_dti': {
        'fallback': lambda bank, tier: TIER_DEFAULTS[tier]['max_dti'],
        'bounds': (20, 70),
        'pick': max,
        'patterns': [
            r'(?:max(?:imum)?\s*)?(?:dti|debt\s*to\s*income)[^\d]{0,18}(\d{1,2}(?:\.\d+)?)\s*%',
        ],
    },
    'max_term_months': {
        'fallback': lambda bank, tier: TIER_DEFAULTS[tier]['max_term_months'],
        'bounds': (24, 96),
        'pick': max,
        'patterns': [
            r'(?:max(?:imum)?\s*)?(?:term|months?|maturity)[^\d]{0,18}(\d{2,3})\s*(?:months?|mos?)',
        ],
    },
    'max_derogatories': {
        'fallback': lambda bank, tier: TIER_DEFAULTS[tier]['max_derogatories'],
        'bounds': (0, 12),
        'pick': max,
        'patterns': [
            r'(?:max(?:imum)?\s*)?(?:derogator(?:y|ies)|charge[-\s]?offs?|collections?)[^\d]{0,18}(\d{1,2})',
        ],
    },
    'max_utilization': {
        'fallback': lambda bank, tier: TIER_DEFAULTS[tier]['max_utilization'],
        'bounds': (10, 100),
        'pick': max,
        'patterns': [
            r'(?:max(?:imum)?\s*)?(?:utilization|revolving\s+utilization)[^\d]{0,18}(\d{1,3}(?:\.\d+)?)\s*%',
        ],
    },
}


@dataclass
class DecodeContext:
    session: requests.Session
    cache_dir: Path
    max_link_depth: int
    max_links_per_resource: int
    timeout_seconds: int
    visited_urls: set[str]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha1_text(value: str) -> str:
    return hashlib.sha1(value.encode('utf-8', errors='ignore')).hexdigest()


def clean_whitespace(text: str) -> str:
    return re.sub(r'\s+', ' ', text or '').strip()


def code_for_bank(bank: str) -> str:
    parts = re.split(r'[^A-Za-z0-9]+', bank)
    if len(parts) == 1:
        return re.sub(r'[^A-Za-z0-9]', '', parts[0]).upper()[:16]
    return ''.join(part[:4].upper() for part in parts if part)[:16]


def discover_bank_docs(bank_root: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not bank_root.exists():
        return rows
    for folder in sorted([item for item in bank_root.iterdir() if item.is_dir()], key=lambda item: item.name.lower()):
        for path in sorted(folder.rglob('*')):
            if not path.is_file():
                continue
            if path.suffix.lower() not in TEXT_EXTENSIONS:
                continue
            rows.append({'bank': folder.name, 'path': path})
    return rows


def html_links(html_text: str, base_url: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html_text, 'html.parser')
    items: list[dict[str, str]] = []
    seen: set[str] = set()
    for anchor in soup.select('a[href]'):
        href = clean_whitespace(anchor.get('href') or '')
        if not href:
            continue
        absolute = urljoin(base_url, href)
        if absolute in seen:
            continue
        seen.add(absolute)
        items.append({'url': absolute, 'label': clean_whitespace(anchor.get_text(' ', strip=True))})
    return items


def relevant_links(items: list[dict[str, str]], base_host: str, max_items: int) -> list[dict[str, str]]:
    filtered: list[dict[str, str]] = []
    for item in items:
        url = item['url']
        parsed = urlparse(url)
        if parsed.scheme not in {'http', 'https'}:
            continue
        label = item.get('label', '')
        haystack = f"{url} {label}"
        same_host = parsed.netloc == base_host if base_host else False
        if same_host or RELEVANT_LINK_RE.search(haystack):
            filtered.append(item)
        if len(filtered) >= max_items:
            break
    return filtered


def snippet(text: str, start: int, end: int, padding: int = 80) -> str:
    left = max(0, start - padding)
    right = min(len(text), end + padding)
    return clean_whitespace(text[left:right])


def extract_pdf(path: Path) -> tuple[str, list[dict[str, Any]], list[dict[str, Any]]]:
    document = fitz.open(path)
    pages: list[dict[str, Any]] = []
    all_links: list[dict[str, Any]] = []
    full_text_parts: list[str] = []
    try:
        for page_number in range(document.page_count):
            page = document.load_page(page_number)
            page_text = page.get_text('text') or ''
            full_text_parts.append(page_text)
            page_links: list[dict[str, Any]] = []
            for link in page.get_links():
                rect = None
                if link.get('from'):
                    try:
                        rect = fitz.Rect(link['from'])
                    except Exception:
                        rect = None
                label = clean_whitespace(page.get_textbox(rect)) if rect else ''
                item = {
                    'kind': int(link.get('kind', 0)) if isinstance(link.get('kind'), int) else link.get('kind'),
                    'uri': clean_whitespace(str(link.get('uri') or '')),
                    'target_page': (int(link.get('page')) + 1) if isinstance(link.get('page'), int) else None,
                    'label': label,
                    'rect': [rect.x0, rect.y0, rect.x1, rect.y1] if rect else None,
                }
                page_links.append(item)
                all_links.append({'page_number': page_number + 1, **item})
            pages.append(
                {
                    'page_number': page_number + 1,
                    'text': page_text,
                    'links': page_links,
                }
            )
    finally:
        document.close()
    return '\n\n'.join(full_text_parts).strip(), pages, all_links


def extract_text_file(path: Path) -> tuple[str, list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix in {'.xlsx', '.xlsm'}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows: list[str] = []
            for sheet in workbook.worksheets:
                rows.append(f'[Sheet] {sheet.title}')
                for row in sheet.iter_rows(values_only=True):
                    values = [clean_whitespace(str(cell)) for cell in row if cell is not None and clean_whitespace(str(cell))]
                    if values:
                        rows.append(' | '.join(values))
            return '\n'.join(rows), []
        finally:
            workbook.close()
    if suffix == '.docx':
        document = DocxDocument(path)
        paragraphs = [clean_whitespace(paragraph.text) for paragraph in document.paragraphs if clean_whitespace(paragraph.text)]
        return '\n'.join(paragraphs), []
    if suffix == '.csv':
        rows: list[str] = []
        with path.open('r', encoding='utf-8', errors='ignore', newline='') as handle:
            reader = csv.reader(handle)
            for row in reader:
                rows.append(', '.join(item.strip() for item in row if item and item.strip()))
        return '\n'.join(rows), []
    raw = path.read_text(encoding='utf-8', errors='ignore')
    if suffix in {'.html', '.htm'}:
        text = trafilatura.extract(raw, include_links=True, include_tables=True) or BeautifulSoup(raw, 'html.parser').get_text(' ', strip=True)
        return text, html_links(raw, path.as_uri())
    if suffix == '.json':
        try:
            parsed = json.loads(raw)
            pretty = json.dumps(parsed, indent=2)
            return pretty, []
        except Exception:
            return raw, []
    return raw, []


def fetch_remote_resource(url: str, bank: str, ctx: DecodeContext, depth: int) -> dict[str, Any] | None:
    if depth > ctx.max_link_depth or url in ctx.visited_urls:
        return None
    ctx.visited_urls.add(url)
    base_host = urlparse(url).netloc
    try:
        response = ctx.session.get(url, timeout=ctx.timeout_seconds, allow_redirects=True)
    except Exception as exc:
        return {'url': url, 'type': 'error', 'error': str(exc), 'depth': depth}

    content_type = clean_whitespace(response.headers.get('content-type') or '').lower()
    final_url = response.url
    if response.status_code >= 400:
        return {'url': url, 'final_url': final_url, 'type': 'error', 'status_code': response.status_code, 'depth': depth}

    if 'pdf' in content_type or final_url.lower().endswith('.pdf'):
        file_name = sanitize_filename(Path(urlparse(final_url).path).name or f'{sha1_text(final_url)}.pdf', 'linked_document.pdf')
        if not file_name.lower().endswith('.pdf'):
            file_name = f'{file_name}.pdf'
        cache_file = ctx.cache_dir / bank / file_name
        cache_file.parent.mkdir(parents=True, exist_ok=True)
        cache_file.write_bytes(response.content)
        text, pages, links = extract_pdf(cache_file)
        nested: list[dict[str, Any]] = []
        if depth < ctx.max_link_depth:
            child_links = [item for item in links if item.get('uri')]
            for child in child_links[: ctx.max_links_per_resource]:
                fetched = fetch_remote_resource(str(child['uri']), bank, ctx, depth + 1)
                if fetched:
                    nested.append(fetched)
        return {
            'url': url,
            'final_url': final_url,
            'type': 'pdf',
            'depth': depth,
            'cached_path': rel(cache_file),
            'text': text,
            'pages': pages,
            'links': links,
            'linked_resources': nested,
        }

    text = response.text
    extracted_text = trafilatura.extract(text, include_links=True, include_tables=True) or BeautifulSoup(text, 'html.parser').get_text(' ', strip=True)
    links = html_links(text, final_url)
    nested_resources: list[dict[str, Any]] = []
    if depth < ctx.max_link_depth:
        for child in relevant_links(links, base_host, ctx.max_links_per_resource):
            fetched = fetch_remote_resource(child['url'], bank, ctx, depth + 1)
            if fetched:
                nested_resources.append(fetched)
    return {
        'url': url,
        'final_url': final_url,
        'type': 'html',
        'depth': depth,
        'title': clean_whitespace(BeautifulSoup(text, 'html.parser').title.get_text()) if BeautifulSoup(text, 'html.parser').title else '',
        'text': extracted_text,
        'links': links,
        'linked_resources': nested_resources,
    }


def decode_document(path: Path, bank: str, ctx: DecodeContext) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == '.pdf':
        text, pages, links = extract_pdf(path)
        remote_resources: list[dict[str, Any]] = []
        for link in [item for item in links if item.get('uri')][: ctx.max_links_per_resource]:
            fetched = fetch_remote_resource(str(link['uri']), bank, ctx, depth=1)
            if fetched:
                remote_resources.append(fetched)
        return {
            'bank': bank,
            'path': rel(path),
            'filename': path.name,
            'type': 'pdf',
            'text': text,
            'pages': pages,
            'links': links,
            'linked_resources': remote_resources,
        }
    text, links = extract_text_file(path)
    remote_resources: list[dict[str, Any]] = []
    base_host = urlparse(path.as_uri()).netloc
    for link in relevant_links(links, base_host, ctx.max_links_per_resource):
        fetched = fetch_remote_resource(link['url'], bank, ctx, depth=1)
        if fetched:
            remote_resources.append(fetched)
    return {
        'bank': bank,
        'path': rel(path),
        'filename': path.name,
        'type': suffix.lstrip('.'),
        'text': text,
        'pages': [],
        'links': links,
        'linked_resources': remote_resources,
    }


def flatten_remote_text(resource: dict[str, Any]) -> list[str]:
    items: list[str] = []
    text = clean_whitespace(str(resource.get('text') or ''))
    if text:
        items.append(text)
    for child in resource.get('linked_resources') or []:
        if isinstance(child, dict):
            items.extend(flatten_remote_text(child))
    return items


def flatten_remote_links(resource: dict[str, Any]) -> list[str]:
    urls: list[str] = []
    url = clean_whitespace(str(resource.get('final_url') or resource.get('url') or ''))
    if url:
        urls.append(url)
    for link in resource.get('links') or []:
        if isinstance(link, dict) and link.get('url'):
            urls.append(str(link['url']))
    for child in resource.get('linked_resources') or []:
        if isinstance(child, dict):
            urls.extend(flatten_remote_links(child))
    return urls


def collect_metric_candidates(text: str, rule: dict[str, Any]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    lower, upper = rule['bounds']
    for pattern in rule['patterns']:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            try:
                value = float(match.group(1))
            except Exception:
                continue
            if value < lower or value > upper:
                continue
            items.append(
                {
                    'value': value,
                    'context': snippet(text, match.start(), match.end()),
                }
            )
    return items


def collect_lines(text: str, keywords: list[str], limit: int = 25) -> list[str]:
    results: list[str] = []
    seen: set[str] = set()
    for raw_line in re.split(r'[\r\n]+', text):
        line = clean_whitespace(raw_line)
        if not line:
            continue
        haystack = line.lower()
        if any(keyword in haystack for keyword in keywords):
            if line not in seen:
                seen.add(line)
                results.append(line)
        if len(results) >= limit:
            break
    return results


def choose_metric(bank: str, tier: str, rule_name: str, text: str) -> tuple[Any, list[dict[str, Any]]]:
    rule = METRIC_RULES[rule_name]
    candidates = collect_metric_candidates(text, rule)
    if candidates:
        chosen = rule['pick'](candidate['value'] for candidate in candidates)
        if float(chosen).is_integer():
            return int(chosen), candidates[:8]
        return round(float(chosen), 2), candidates[:8]
    return rule['fallback'](bank, tier), []


def build_profile(bank: str, bank_docs: list[dict[str, Any]]) -> dict[str, Any]:
    tier = BANK_TIER_HINTS.get(bank, 'near_prime')
    defaults = TIER_DEFAULTS[tier]
    local_text = '\n\n'.join(clean_whitespace(doc.get('text', '')) for doc in bank_docs if doc.get('text'))
    remote_text = '\n\n'.join(
        text for doc in bank_docs for resource in doc.get('linked_resources') or [] for text in flatten_remote_text(resource)
    )
    aggregate_text = '\n\n'.join(part for part in [local_text, remote_text] if part)

    metrics: dict[str, Any] = {}
    evidence: dict[str, list[dict[str, Any]]] = {}
    for metric_name in METRIC_RULES:
        value, metric_evidence = choose_metric(bank, tier, metric_name, aggregate_text)
        metrics[metric_name] = value
        evidence[metric_name] = metric_evidence

    stips = collect_lines(aggregate_text, STIP_KEYWORDS)
    restrictions = collect_lines(aggregate_text, RULE_KEYWORDS)
    source_paths = [doc['path'] for doc in bank_docs]
    source_links = sorted({url for doc in bank_docs for resource in doc.get('linked_resources') or [] for url in flatten_remote_links(resource) if url})
    evidence_count = sum(len(items) for items in evidence.values()) + len(stips) + len(restrictions)
    confidence = min(0.96, 0.18 + (0.06 * len(bank_docs)) + (0.015 * len(source_links)) + (0.025 * evidence_count))

    notes: list[str] = []
    if not bank_docs:
        notes.append('No bank docs decoded yet. Using lender-tier fallback defaults until RouteOne docs are imported.')
    elif evidence_count == 0:
        notes.append('Docs were decoded but few numeric rule anchors were detected. Default tier limits were retained.')
    if source_links:
        notes.append(f'Linked content inspected: {len(source_links)} unique URLs.')

    return {
        'code': code_for_bank(bank),
        'bank': bank.replace('_', ' '),
        'name': bank.replace('_', ' '),
        'bank_folder': bank,
        'tier': tier,
        'weight': round(max(defaults['weight'], defaults['weight'] + confidence - 0.35), 2),
        'confidence': round(confidence * 100, 1),
        'min_score': metrics['min_score'],
        'max_ltv': metrics['max_ltv'],
        'max_pti': metrics['max_pti'],
        'max_dti': metrics['max_dti'],
        'max_derogatories': metrics['max_derogatories'],
        'max_utilization': metrics['max_utilization'],
        'max_term_months': metrics['max_term_months'],
        'source_files': source_paths,
        'source_links': source_links[:100],
        'stips': stips[:25],
        'restrictions': restrictions[:25],
        'notes': notes,
        'rule_evidence': {key: value[:6] for key, value in evidence.items() if value},
        'decoded_doc_count': len(bank_docs),
        'decoded_link_count': len(source_links),
        'decoded_evidence_count': evidence_count,
    }


def rebuild_bank_brain(args: argparse.Namespace) -> dict[str, Any]:
    bank_root = Path(args.bank_root).resolve()
    decoded_dir = Path(args.decoded_dir).resolve()
    index_path = Path(args.index_path).resolve()
    profiles_path = Path(args.profiles_path).resolve()
    sales_banks_path = Path(args.sales_banks_path).resolve()
    cache_dir = Path(args.link_cache_dir).resolve()

    decoded_dir.mkdir(parents=True, exist_ok=True)
    cache_dir.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers.update({'User-Agent': 'RouteOneBankBrainRebuild/1.0'})
    ctx = DecodeContext(
        session=session,
        cache_dir=cache_dir,
        max_link_depth=args.max_link_depth,
        max_links_per_resource=args.max_links_per_resource,
        timeout_seconds=args.timeout_seconds,
        visited_urls=set(),
    )

    aliases = build_bank_aliases(bank_root)
    docs = discover_bank_docs(bank_root)
    decoded_docs: list[dict[str, Any]] = []
    errors: list[str] = []

    for entry in docs:
        path = entry['path']
        bank = entry['bank']
        if bank not in aliases:
            bank = infer_bank(path.name, aliases)
        try:
            decoded = decode_document(path, bank, ctx)
            decoded_docs.append(decoded)
            bank_dir = decoded_dir / bank
            bank_dir.mkdir(parents=True, exist_ok=True)
            out_file = bank_dir / f"{sanitize_filename(path.stem, 'document')}.json"
            out_file.write_text(json.dumps(decoded, indent=2), encoding='utf-8')
        except Exception as exc:
            errors.append(f'{path}: {exc}')

    by_bank: dict[str, list[dict[str, Any]]] = {bank: [] for bank in aliases}
    for decoded in decoded_docs:
        by_bank.setdefault(str(decoded['bank']), []).append(decoded)

    profiles = [build_profile(bank, by_bank.get(bank, [])) for bank in sorted(aliases)]

    index_payload = {
        'generated_at': now_iso(),
        'bank_root': rel(bank_root),
        'decoded_dir': rel(decoded_dir),
        'linked_cache_dir': rel(cache_dir),
        'doc_count': len(decoded_docs),
        'bank_count': len(profiles),
        'errors': errors,
        'documents': [
            {
                'bank': doc['bank'],
                'path': doc['path'],
                'type': doc['type'],
                'page_count': len(doc.get('pages') or []),
                'link_count': len(doc.get('links') or []),
                'linked_resource_count': len(doc.get('linked_resources') or []),
                'text_length': len(doc.get('text') or ''),
            }
            for doc in decoded_docs
        ],
    }
    index_path.parent.mkdir(parents=True, exist_ok=True)
    index_path.write_text(json.dumps(index_payload, indent=2), encoding='utf-8')

    profiles_payload = {
        'generated_at': now_iso(),
        'source_index': rel(index_path),
        'profiles': profiles,
    }
    profiles_path.parent.mkdir(parents=True, exist_ok=True)
    profiles_path.write_text(json.dumps(profiles_payload, indent=2), encoding='utf-8')

    sales_payload = {
        'version': '0.2.0',
        'generatedAt': now_iso(),
        'policies': profiles,
    }
    sales_banks_path.parent.mkdir(parents=True, exist_ok=True)
    sales_banks_path.write_text(json.dumps(sales_payload, indent=2), encoding='utf-8')

    return {
        'ok': not errors,
        'generated_at': now_iso(),
        'decoded_docs': len(decoded_docs),
        'profiles_generated': len(profiles),
        'doc_index_path': rel(index_path),
        'profiles_path': rel(profiles_path),
        'sales_banks_path': rel(sales_banks_path),
        'errors': errors,
    }


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Decode bank docs plus linked content and rebuild lender policy profiles.')
    parser.add_argument('--bank-root', default=str(DEFAULT_BANK_ROOT))
    parser.add_argument('--decoded-dir', default=str(DEFAULT_DECODED_DIR))
    parser.add_argument('--index-path', default=str(DEFAULT_INDEX_PATH))
    parser.add_argument('--profiles-path', default=str(DEFAULT_PROFILES_PATH))
    parser.add_argument('--sales-banks-path', default=str(DEFAULT_SALES_BANKS_PATH))
    parser.add_argument('--link-cache-dir', default=str(DEFAULT_LINK_CACHE_DIR))
    parser.add_argument('--max-link-depth', type=int, default=1)
    parser.add_argument('--max-links-per-resource', type=int, default=12)
    parser.add_argument('--timeout-seconds', type=int, default=20)
    parser.add_argument('--json', action='store_true')
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    payload = rebuild_bank_brain(args)
    if args.json:
        print(json.dumps(payload, indent=2))
    return 0 if payload.get('ok', False) else 2


if __name__ == '__main__':
    raise SystemExit(main(sys.argv[1:]))

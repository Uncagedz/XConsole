from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from export_liftedtrucks_dashboard import _latest_intelligence_csv, _to_float, _to_int


OUTPUT_XLSX = Path(r"D:\liftedtrucks_raw_equipment_20260401.xlsx")


THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)


GROUP_FILLS = {
    "vehicle": PatternFill("solid", fgColor="EDE7D8"),
    "color": PatternFill("solid", fgColor="E8EEF3"),
    "pricing": PatternFill("solid", fgColor="E7F0E8"),
    "inventory": PatternFill("solid", fgColor="E7EEF7"),
    "wheels": PatternFill("solid", fgColor="F2E8F1"),
    "tires": PatternFill("solid", fgColor="EAF4EC"),
    "lift": PatternFill("solid", fgColor="F6EADF"),
    "mods": PatternFill("solid", fgColor="F3EEE4"),
}


def _clean(value: str | None) -> str:
    if value is None:
        return ""
    text = str(value).replace("[INFERRED]", "").strip()
    bad = [
        "unknown",
        "not explicitly listed",
        "not explicitly",
        "none clearly stated",
    ]
    lowered = text.lower()
    if not text or any(fragment in lowered for fragment in bad):
        return ""
    return " ".join(text.split())


def _vehicle_title(row: dict) -> str:
    return " ".join(
        part for part in [
            str(row.get("year") or "").strip(),
            str(row.get("make") or "").strip(),
            str(row.get("model") or "").strip(),
            str(row.get("trim") or "").strip(),
        ]
        if part
    )


def _color_fill(color_name: str) -> tuple[PatternFill, str]:
    name = (color_name or "").lower()
    if not name:
        return PatternFill("solid", fgColor="F8F7F3"), "1F2937"
    mapping = [
        (["black", "diamond black", "jet black"], ("2F343A", "FFFFFF")),
        (["white", "bright white", "ivory"], ("F5F5F1", "1F2937")),
        (["hydro blue", "blue", "navy", "indigo"], ("CFE1F7", "1F2937")),
        (["red", "firecracker", "delmonico", "velvet"], ("F8D7D5", "7A1E1E")),
        (["green", "sarge", "olive"], ("D9E7D4", "244226")),
        (["yellow", "gold", "baja"], ("F5E7A7", "5F4B00")),
        (["orange", "copper"], ("F5D7BD", "7A3E00")),
        (["gray", "grey", "granite", "anvil", "silver", "sting"], ("E3E6EA", "374151")),
        (["tan", "beige", "brown", "mojave"], ("E9DCC7", "5B4630")),
    ]
    for keys, (fill, font) in mapping:
        if any(key in name for key in keys):
            return PatternFill("solid", fgColor=fill), font
    return PatternFill("solid", fgColor="EEF1F4"), "1F2937"


def load_rows() -> list[dict]:
    source_csv = _latest_intelligence_csv()
    with source_csv.open("r", newline="", encoding="utf-8") as handle:
        source_rows = list(csv.DictReader(handle))

    rows: list[dict] = []
    for row in source_rows:
        vehicle_price = _to_float(row.get("vehicle_price_usd"))
        featured_price = _to_float(row.get("featured_price_usd"))
        aftermarket_price = _to_float(row.get("aftermarket_accessories_usd"))
        days_in_stock = _to_int(row.get("days_in_stock"))
        clicks_7d = _to_int(row.get("website_clicks_public"))
        engagement = _to_float(row.get("public_views_per_live_day"))

        rows.append(
            {
                "Vehicle": _vehicle_title(row),
                "Year": _to_int(row.get("year")),
                "Make": row.get("make"),
                "Model": row.get("model"),
                "Trim": row.get("trim"),
                "VIN": row.get("vin"),
                "Exterior Color": _clean(row.get("exterior_color")),
                "Interior Color": _clean(row.get("interior_color")),
                "Vehicle Price": vehicle_price,
                "Featured Price": featured_price,
                "Aftermarket Price": aftermarket_price,
                "Days In Stock": days_in_stock,
                "Clicks / 7d": clicks_7d,
                "Engagement / Day": engagement,
                "Wheel Style": _clean(row.get("wheel_style")),
                "Wheel Size": _clean(row.get("wheel_size_inches")),
                "Tire Model": _clean(row.get("tire_brand_model")),
                "Tire Size": _clean(row.get("tire_size")),
                "Tire Type": _clean(row.get("tire_type")),
                "Lift Kit": _clean(row.get("lift_brand_model")),
                "Lift Type": _clean(row.get("lift_type")),
                "Lift Height (in)": _clean(row.get("lift_height_inches")),
                "Other Mods": _clean(row.get("other_mods")),
                "Build Tier": _clean(row.get("build_quality_tier")),
                "Confidence": row.get("confidence_level"),
            }
        )
    rows.sort(key=lambda item: ((item["Clicks / 7d"] or -1), (item["Engagement / Day"] or -1), -(item["Days In Stock"] or 9999)), reverse=True)
    return rows


def _column_groups() -> dict[str, str]:
    return {
        "Vehicle": "vehicle",
        "Year": "vehicle",
        "Make": "vehicle",
        "Model": "vehicle",
        "Trim": "vehicle",
        "VIN": "vehicle",
        "Exterior Color": "color",
        "Interior Color": "color",
        "Vehicle Price": "pricing",
        "Featured Price": "pricing",
        "Aftermarket Price": "pricing",
        "Days In Stock": "inventory",
        "Clicks / 7d": "inventory",
        "Engagement / Day": "inventory",
        "Wheel Style": "wheels",
        "Wheel Size": "wheels",
        "Tire Model": "tires",
        "Tire Size": "tires",
        "Tire Type": "tires",
        "Lift Kit": "lift",
        "Lift Type": "lift",
        "Lift Height (in)": "lift",
        "Other Mods": "mods",
        "Build Tier": "mods",
        "Confidence": "mods",
    }


def _autosize(ws) -> None:
    widths = {
        "A": 28,
        "B": 8,
        "C": 12,
        "D": 14,
        "E": 18,
        "F": 22,
        "G": 24,
        "H": 22,
        "I": 12,
        "J": 12,
        "K": 14,
        "L": 11,
        "M": 11,
        "N": 13,
        "O": 28,
        "P": 12,
        "Q": 24,
        "R": 14,
        "S": 10,
        "T": 24,
        "U": 14,
        "V": 12,
        "W": 26,
        "X": 18,
        "Y": 11,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_workbook(rows: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Equipment"
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows) + 1}"

    headers = list(rows[0].keys())
    groups = _column_groups()

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        cell.fill = GROUP_FILLS[groups[header]]
    ws.row_dimensions[1].height = 26

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row[header]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=header in {"Vehicle", "Wheel Style", "Tire Model", "Lift Kit", "Other Mods"})
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FCFBF8")

            if header in {"Vehicle Price", "Featured Price", "Aftermarket Price"} and value is not None:
                cell.number_format = "$#,##0"
            if header == "Engagement / Day" and value is not None:
                cell.number_format = "0.00"
            if header in {"Year", "Days In Stock", "Clicks / 7d"} and value is not None:
                cell.number_format = "0"

        exterior = ws.cell(row=row_idx, column=headers.index("Exterior Color") + 1)
        fill, font_color = _color_fill(str(row["Exterior Color"]))
        exterior.fill = fill
        exterior.font = Font(color=font_color, bold=bool(row["Exterior Color"]))

    _autosize(ws)

    days_col = get_column_letter(headers.index("Days In Stock") + 1)
    clicks_col = get_column_letter(headers.index("Clicks / 7d") + 1)
    engagement_col = get_column_letter(headers.index("Engagement / Day") + 1)

    ws.conditional_formatting.add(
        f"{days_col}2:{days_col}{len(rows)+1}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="D9F0D8",
            mid_type="num", mid_value=45, mid_color="F8E59A",
            end_type="num", end_value=180, end_color="F4B4AE",
        ),
    )
    ws.conditional_formatting.add(
        f"{clicks_col}2:{clicks_col}{len(rows)+1}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="EDF4FF",
            mid_type="num", mid_value=10, mid_color="B7D4F9",
            end_type="num", end_value=45, end_color="5A98E6",
        ),
    )
    ws.conditional_formatting.add(
        f"{engagement_col}2:{engagement_col}{len(rows)+1}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="EAF7EE",
            mid_type="num", mid_value=1.5, mid_color="BDE3C7",
            end_type="num", end_value=6.5, end_color="4E9B63",
        ),
    )

    return wb


def main() -> int:
    rows = load_rows()
    wb = build_workbook(rows)
    wb.save(OUTPUT_XLSX)
    print(json.dumps({"ok": True, "xlsx": str(OUTPUT_XLSX), "row_count": len(rows)}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

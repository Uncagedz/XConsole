from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

SALESPEOPLE = [
    "Alex",
    "Andres",
    "Ani",
    "Arturo",
    "Austin",
    "Azlin",
    "Brandon",
    "Brandon Budhai",
    "Chris",
    "Clara",
    "Darren",
    "Eric",
    "Gregory",
    "Joel",
    "Jonathan",
    "Kenice",
    "Marlon",
    "Michael",
    "Mikey",
    "Pedro",
    "Reynel",
    "Rollin",
    "Xavier",
    "Zachary",
]

SOFTWARES = [
    "Outlook",
    "DriveCentric",
    "Reconversion",
    "One Micro",
    "AccuTrade",
    "monday.com",
    "Dealer Connect",
]

SOFTWARE_STATUS_SCORES = [
    ("Not Started", 0.00),
    ("Requested", 0.15),
    ("Provisioning", 0.35),
    ("Password Reset Needed", 0.45),
    ("Active", 0.70),
    ("Verified", 1.00),
    ("Training Needed", 0.80),
    ("Blocked", 0.10),
]

WORKFLOW_STATUS_SCORES = [
    ("Not Started", 0.00),
    ("Scheduled", 0.50),
    ("Completed", 1.00),
    ("Needs Coaching", 0.60),
]

SIGNOFF_STATUS_SCORES = [
    ("Pending", 0.00),
    ("Approved", 1.00),
    ("Hold", 0.25),
]

REVIEW_TYPES = [
    "Current Team Audit",
    "New Hire Onboarding",
    "Quarterly Recheck",
]

STATUS_COLORS = {
    "Not Started": "D9D9D9",
    "Requested": "FFF2CC",
    "Provisioning": "FFE599",
    "Password Reset Needed": "F9CB9C",
    "Active": "CFE2F3",
    "Verified": "D9EAD3",
    "Training Needed": "FCE5CD",
    "Blocked": "EA9999",
    "Scheduled": "FFF2CC",
    "Completed": "D9EAD3",
    "Needs Coaching": "FCE5CD",
    "Pending": "FFF2CC",
    "Approved": "D9EAD3",
    "Hold": "EA9999",
    "Access In Progress": "FFF2CC",
    "Access Live - Verify": "CFE2F3",
    "Needs Training": "FCE5CD",
    "Pending Final Review": "D9D2E9",
    "Ready To Sell": "D9EAD3",
    "Urgent": "EA9999",
    "High": "FCE5CD",
    "Medium": "CFE2F3",
    "Low": "D9D9D9",
    "Done": "D9EAD3",
}

DARK_BLUE = "17375E"
MID_BLUE = "DCE6F1"
LIGHT_BLUE = "EAF2F8"
SOFT_YELLOW = "FFFDF2"
SOFT_GREEN = "EEF7EC"
SOFT_RED = "FDECEC"
SOFT_PURPLE = "F3ECFB"
SECTION_PEOPLE = "D9E2F3"
SECTION_AUTO = "DCE6F1"
SECTION_ACTIONS = "EADCF8"
SECTION_SOFTWARE = "D9EAD3"
WHITE = "FFFFFF"
BLACK = "000000"
GRAY_TEXT = "555555"

THIN = Side(style="thin", color="C9D1D9")
MEDIUM = Side(style="medium", color=DARK_BLUE)
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "docs" / "sales-access-control-center.xlsx"


def fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def style_cell(
    cell,
    *,
    fill_color: str | None = None,
    font_size: int = 11,
    bold: bool = False,
    font_color: str = BLACK,
    h_align: str = "left",
    v_align: str = "center",
    wrap: bool = False,
    border: Border | None = THIN_BORDER,
    number_format: str | None = None,
) -> None:
    cell.font = Font(name="Calibri", size=font_size, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    if fill_color:
        cell.fill = fill(fill_color)
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def merge_label(ws, cell_range: str, value: str, bg: str, *, size: int = 12, color: str = WHITE) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    style_cell(cell, fill_color=bg, font_size=size, bold=True, font_color=color, h_align="center")


def add_named_range(workbook: Workbook, name: str, ref: str) -> None:
    workbook.defined_names.add(DefinedName(name, attr_text=ref))


def add_text_fill_rules(ws, cell_range: str, base_cell: str, values: list[str]) -> None:
    for value in values:
        color = STATUS_COLORS[value]
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'{base_cell}="{value}"'], fill=fill(color)),
        )


def build_lists_sheet(wb: Workbook):
    ws = wb.create_sheet("Lists")
    ws.sheet_state = "hidden"

    ws["A1"] = "Software Status"
    ws["B1"] = "Score"
    for row, (label, score) in enumerate(SOFTWARE_STATUS_SCORES, start=2):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=score)

    ws["D1"] = "Workflow Demo"
    ws["E1"] = "Score"
    for row, (label, score) in enumerate(WORKFLOW_STATUS_SCORES, start=2):
        ws.cell(row=row, column=4, value=label)
        ws.cell(row=row, column=5, value=score)

    ws["G1"] = "Manager Sign-Off"
    ws["H1"] = "Score"
    for row, (label, score) in enumerate(SIGNOFF_STATUS_SCORES, start=2):
        ws.cell(row=row, column=7, value=label)
        ws.cell(row=row, column=8, value=score)

    ws["J1"] = "Review Type"
    for row, label in enumerate(REVIEW_TYPES, start=2):
        ws.cell(row=row, column=10, value=label)

    add_named_range(wb, "SoftwareStatusList", "Lists!$A$2:$A$9")
    add_named_range(wb, "SoftwareStatusLabels", "Lists!$A$2:$A$9")
    add_named_range(wb, "SoftwareStatusScores", "Lists!$B$2:$B$9")
    add_named_range(wb, "WorkflowDemoList", "Lists!$D$2:$D$5")
    add_named_range(wb, "WorkflowLabels", "Lists!$D$2:$D$5")
    add_named_range(wb, "WorkflowScores", "Lists!$E$2:$E$5")
    add_named_range(wb, "SignoffList", "Lists!$G$2:$G$4")
    add_named_range(wb, "SignoffLabels", "Lists!$G$2:$G$4")
    add_named_range(wb, "SignoffScores", "Lists!$H$2:$H$4")
    add_named_range(wb, "ReviewTypeList", "Lists!$J$2:$J$4")

    return ws


def build_tracker_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Tracker"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DARK_BLUE
    ws.freeze_panes = "F5"

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 34

    merge_label(ws, "A1:W1", "Sales Access Control Center", DARK_BLUE, size=16)
    merge_label(
        ws,
        "A2:W2",
        f"Use the dropdowns to track access, live verification, coaching, and final readiness. Generated {date.today():%B %d, %Y}.",
        MID_BLUE,
        size=11,
        color=DARK_BLUE,
    )
    merge_label(ws, "A3:D3", "People & Ownership", SECTION_PEOPLE, size=11, color=DARK_BLUE)
    merge_label(ws, "E3:J3", "Auto Calculated", SECTION_AUTO, size=11, color=DARK_BLUE)
    merge_label(ws, "K3:P3", "Manager Actions", SECTION_ACTIONS, size=11, color=DARK_BLUE)
    merge_label(ws, "Q3:W3", "Required Software", SECTION_SOFTWARE, size=11, color=DARK_BLUE)

    headers = [
        "Salesperson",
        "Department",
        "Review Type",
        "Access Owner",
        "Overall Status",
        "Priority",
        "Completion %",
        "Verified Apps",
        "Apps Missing Verification",
        "Live Access Count",
        "Workflow Demo",
        "Manager Sign-Off",
        "Primary Blocker",
        "Last Reviewed",
        "Next Follow-Up",
        "Notes",
        *SOFTWARES,
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        style_cell(
            cell,
            fill_color=DARK_BLUE,
            font_size=10,
            bold=True,
            font_color=WHITE,
            h_align="center",
            wrap=True,
            border=Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM),
        )

    comments = {
        "E4": "Auto-calculated from the seven app statuses, workflow demo status, and manager sign-off.",
        "F4": "Auto-calculated to help you sort the team by urgency.",
        "G4": "Weighted progress score across all required apps plus workflow and final approval.",
        "H4": "Count of required apps marked Verified.",
        "I4": "How many required apps still need to reach Verified.",
        "J4": "Count of app statuses that are Active, Verified, or Training Needed.",
        "K4": "Use after the salesperson demonstrates the real daily workflow.",
        "L4": "Final readiness decision from management.",
        "M4": "Short plain-English blocker note so anyone can pick this up.",
    }
    for column in range(17, 24):
        comments[f"{get_column_letter(column)}4"] = "Use the dropdown. Verified means the salesperson personally logged in and confirmed they can use this system."
    for address, text in comments.items():
        ws[address].comment = Comment(text, "Codex")

    widths = {
        "A": 20,
        "B": 12,
        "C": 20,
        "D": 18,
        "E": 22,
        "F": 12,
        "G": 13,
        "H": 12,
        "I": 22,
        "J": 14,
        "K": 18,
        "L": 18,
        "M": 24,
        "N": 14,
        "O": 15,
        "P": 30,
        "Q": 16,
        "R": 16,
        "S": 16,
        "T": 16,
        "U": 16,
        "V": 16,
        "W": 16,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    start_row = 5
    for row_idx, salesperson in enumerate(SALESPEOPLE, start=start_row):
        ws.row_dimensions[row_idx].height = 22
        overall_formula = (
            f'=IF(A{row_idx}="","",'
            f'IF(OR(COUNTIF(Q{row_idx}:W{row_idx},"Blocked")>0,L{row_idx}="Hold"),"Blocked",'
            f'IF(AND(COUNTIF(Q{row_idx}:W{row_idx},"Verified")=7,K{row_idx}="Completed",L{row_idx}="Approved"),"Ready To Sell",'
            f'IF(OR(COUNTIF(Q{row_idx}:W{row_idx},"Training Needed")>0,K{row_idx}="Needs Coaching"),"Needs Training",'
            f'IF(AND(COUNTIF(Q{row_idx}:W{row_idx},"Verified")=7,OR(K{row_idx}<>"Completed",L{row_idx}<>"Approved")),"Pending Final Review",'
            f'IF(COUNTIF(Q{row_idx}:W{row_idx},"Active")+COUNTIF(Q{row_idx}:W{row_idx},"Verified")+COUNTIF(Q{row_idx}:W{row_idx},"Training Needed")>0,"Access Live - Verify",'
            f'IF(COUNTIF(Q{row_idx}:W{row_idx},"Requested")+COUNTIF(Q{row_idx}:W{row_idx},"Provisioning")+COUNTIF(Q{row_idx}:W{row_idx},"Password Reset Needed")>0,"Access In Progress","Not Started"))))))'
        )
        priority_formula = (
            f'=IF(A{row_idx}="","",'
            f'IF(E{row_idx}="Blocked","Urgent",'
            f'IF(OR(E{row_idx}="Needs Training",E{row_idx}="Access In Progress",E{row_idx}="Pending Final Review"),"High",'
            f'IF(E{row_idx}="Access Live - Verify","Medium",'
            f'IF(E{row_idx}="Ready To Sell","Done","Low")))))'
        )
        completion_formula = (
            f'=IF(A{row_idx}="","",('
            f'IFERROR(VLOOKUP(Q{row_idx},Lists!$A$2:$B$9,2,FALSE),0)+'
            f'IFERROR(VLOOKUP(R{row_idx},Lists!$A$2:$B$9,2,FALSE),0)+'
            f'IFERROR(VLOOKUP(S{row_idx},Lists!$A$2:$B$9,2,FALSE),0)+'
            f'IFERROR(VLOOKUP(T{row_idx},Lists!$A$2:$B$9,2,FALSE),0)+'
            f'IFERROR(VLOOKUP(U{row_idx},Lists!$A$2:$B$9,2,FALSE),0)+'
            f'IFERROR(VLOOKUP(V{row_idx},Lists!$A$2:$B$9,2,FALSE),0)+'
            f'IFERROR(VLOOKUP(W{row_idx},Lists!$A$2:$B$9,2,FALSE),0)+'
            f'IFERROR(VLOOKUP(K{row_idx},Lists!$D$2:$E$5,2,FALSE),0)+'
            f'IFERROR(VLOOKUP(L{row_idx},Lists!$G$2:$H$4,2,FALSE),0)'
            f')/9)'
        )
        verified_formula = f'=IF(A{row_idx}="","",COUNTIF(Q{row_idx}:W{row_idx},"Verified"))'
        missing_formula = f'=IF(A{row_idx}="","",7-H{row_idx})'
        live_formula = (
            f'=IF(A{row_idx}="","",'
            f'COUNTIF(Q{row_idx}:W{row_idx},"Active")+COUNTIF(Q{row_idx}:W{row_idx},"Verified")+COUNTIF(Q{row_idx}:W{row_idx},"Training Needed"))'
        )
        row_values = [
            salesperson,
            "Sales",
            "Current Team Audit",
            "",
            overall_formula,
            priority_formula,
            completion_formula,
            verified_formula,
            missing_formula,
            live_formula,
            "Not Started",
            "Pending",
            "",
            "",
            "",
            "",
            *["Not Started"] * len(SOFTWARES),
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if 5 <= col_idx <= 10:
                style_cell(cell, fill_color=LIGHT_BLUE, font_color=DARK_BLUE)
            elif col_idx in {11, 12, 13, 14, 15, 16}:
                style_cell(cell, fill_color=SOFT_YELLOW, wrap=col_idx in {13, 16})
            elif col_idx >= 17:
                style_cell(cell, fill_color=WHITE, h_align="center")
            else:
                style_cell(cell)

        ws.cell(row=row_idx, column=7).number_format = "0%"
        ws.cell(row=row_idx, column=14).number_format = "mmm d, yyyy"
        ws.cell(row=row_idx, column=15).number_format = "mmm d, yyyy"

    software_dv = DataValidation(type="list", formula1="=SoftwareStatusList", allow_blank=True)
    software_dv.promptTitle = "Software Status"
    software_dv.prompt = "Pick the access status for this software."
    software_dv.errorTitle = "Use the dropdown"
    software_dv.error = "Choose one of the approved software status values."
    ws.add_data_validation(software_dv)
    software_dv.add("Q5:W200")

    workflow_dv = DataValidation(type="list", formula1="=WorkflowDemoList", allow_blank=True)
    workflow_dv.promptTitle = "Workflow Demo"
    workflow_dv.prompt = "Track whether the salesperson has completed the live workflow demo."
    ws.add_data_validation(workflow_dv)
    workflow_dv.add("K5:K200")

    signoff_dv = DataValidation(type="list", formula1="=SignoffList", allow_blank=True)
    signoff_dv.promptTitle = "Manager Sign-Off"
    signoff_dv.prompt = "Use Approved only when the salesperson is truly ready to sell."
    ws.add_data_validation(signoff_dv)
    signoff_dv.add("L5:L200")

    review_dv = DataValidation(type="list", formula1="=ReviewTypeList", allow_blank=True)
    review_dv.promptTitle = "Review Type"
    review_dv.prompt = "Pick the context for this review."
    ws.add_data_validation(review_dv)
    review_dv.add("C5:C200")

    date_dv = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2025,1,1)",
        formula2="DATE(2035,12,31)",
        allow_blank=True,
    )
    date_dv.promptTitle = "Date"
    date_dv.prompt = "Use a valid calendar date."
    ws.add_data_validation(date_dv)
    date_dv.add("N5:O200")

    add_text_fill_rules(ws, "Q5:W200", "Q5", [label for label, _ in SOFTWARE_STATUS_SCORES])
    add_text_fill_rules(
        ws,
        "E5:E200",
        "E5",
        [
            "Not Started",
            "Access In Progress",
            "Access Live - Verify",
            "Needs Training",
            "Pending Final Review",
            "Ready To Sell",
            "Blocked",
        ],
    )
    add_text_fill_rules(ws, "F5:F200", "F5", ["Low", "Medium", "High", "Urgent", "Done"])
    add_text_fill_rules(ws, "K5:K200", "K5", [label for label, _ in WORKFLOW_STATUS_SCORES])
    add_text_fill_rules(ws, "L5:L200", "L5", [label for label, _ in SIGNOFF_STATUS_SCORES])

    ws.conditional_formatting.add(
        "A5:P200",
        FormulaRule(formula=['$E5="Blocked"'], fill=fill(SOFT_RED)),
    )
    ws.conditional_formatting.add(
        "A5:P200",
        FormulaRule(formula=['$E5="Ready To Sell"'], fill=fill(SOFT_GREEN)),
    )
    ws.conditional_formatting.add(
        "A5:P200",
        FormulaRule(formula=['$E5="Pending Final Review"'], fill=fill(SOFT_PURPLE)),
    )
    ws.conditional_formatting.add(
        "O5:O200",
        FormulaRule(
            formula=['AND($O5<>"",$O5<TODAY(),$E5<>"Ready To Sell")'],
            fill=fill("F4CCCC"),
        ),
    )
    ws.conditional_formatting.add(
        "M5:M200",
        FormulaRule(
            formula=['AND($M5<>"",$E5<>"Ready To Sell")'],
            fill=fill("FCE5CD"),
        ),
    )
    ws.conditional_formatting.add(
        "G5:G200",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63C384"),
    )

    ws.auto_filter.ref = f"A4:W{start_row + len(SALESPEOPLE) - 1}"

    return ws


def build_dashboard_sheet(wb: Workbook):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "4F81BD"

    merge_label(ws, "A1:L1", "Sales Readiness Dashboard", DARK_BLUE, size=16)
    merge_label(
        ws,
        "A2:L2",
        "Filter and update the Tracker sheet, then use this page for the management view.",
        MID_BLUE,
        size=11,
        color=DARK_BLUE,
    )

    for column in range(1, 13):
        ws.column_dimensions[get_column_letter(column)].width = 14

    def kpi(top_row: int, left_col: int, title: str, formula: str, color: str, numfmt: str = "0") -> None:
        ws.merge_cells(start_row=top_row, start_column=left_col, end_row=top_row, end_column=left_col + 1)
        ws.merge_cells(start_row=top_row + 1, start_column=left_col, end_row=top_row + 2, end_column=left_col + 1)
        label = ws.cell(row=top_row, column=left_col, value=title)
        value = ws.cell(row=top_row + 1, column=left_col, value=formula)
        style_cell(label, fill_color=DARK_BLUE, bold=True, font_color=WHITE, h_align="center")
        style_cell(value, fill_color=color, bold=True, font_size=20, h_align="center", border=THIN_BORDER, number_format=numfmt)
        for merged_col in (left_col + 1,):
            style_cell(ws.cell(row=top_row + 1, column=merged_col), fill_color=color, border=THIN_BORDER)

    kpi(4, 2, "Total Team", '=COUNTA(Tracker!$A$5:$A$200)', MID_BLUE)
    kpi(4, 5, "Ready To Sell", '=COUNTIF(Tracker!$E$5:$E$200,"Ready To Sell")', STATUS_COLORS["Ready To Sell"])
    kpi(4, 8, "Open Items", '=COUNTA(Tracker!$A$5:$A$200)-COUNTIF(Tracker!$E$5:$E$200,"Ready To Sell")', STATUS_COLORS["Pending Final Review"])
    kpi(4, 11, "Urgent", '=COUNTIF(Tracker!$F$5:$F$200,"Urgent")', STATUS_COLORS["Urgent"])
    kpi(8, 2, "Avg Completion", '=AVERAGEIF(Tracker!$A$5:$A$200,"<>",Tracker!$G$5:$G$200)', STATUS_COLORS["Access Live - Verify"], "0%")
    kpi(8, 5, "Needs Training", '=COUNTIF(Tracker!$E$5:$E$200,"Needs Training")', STATUS_COLORS["Needs Training"])
    kpi(8, 8, "Access In Progress", '=COUNTIF(Tracker!$E$5:$E$200,"Access In Progress")', STATUS_COLORS["Access In Progress"])
    kpi(8, 11, "Pending Final Review", '=COUNTIF(Tracker!$E$5:$E$200,"Pending Final Review")', STATUS_COLORS["Pending Final Review"])

    ws["A13"] = "Status Mix"
    style_cell(ws["A13"], fill_color=DARK_BLUE, bold=True, font_color=WHITE, h_align="center")
    ws["B13"] = "Count"
    style_cell(ws["B13"], fill_color=DARK_BLUE, bold=True, font_color=WHITE, h_align="center")

    status_rows = [
        "Ready To Sell",
        "Pending Final Review",
        "Needs Training",
        "Access Live - Verify",
        "Access In Progress",
        "Blocked",
        "Not Started",
    ]
    for row_idx, status in enumerate(status_rows, start=14):
        ws.cell(row=row_idx, column=1, value=status)
        ws.cell(row=row_idx, column=2, value=f'=COUNTIF(Tracker!$E$5:$E$200,A{row_idx})')
        style_cell(ws.cell(row=row_idx, column=1), fill_color=STATUS_COLORS[status], bold=True)
        style_cell(ws.cell(row=row_idx, column=2), h_align="center")

    ws["D13"] = "App Gap Analysis"
    style_cell(ws["D13"], fill_color=DARK_BLUE, bold=True, font_color=WHITE, h_align="center")
    ws["E13"] = "People Not Yet Verified"
    style_cell(ws["E13"], fill_color=DARK_BLUE, bold=True, font_color=WHITE, h_align="center")

    for row_idx, app in enumerate(SOFTWARES, start=14):
        col_letter = get_column_letter(17 + SOFTWARES.index(app))
        ws.cell(row=row_idx, column=4, value=app)
        ws.cell(
            row=row_idx,
            column=5,
            value=f'=COUNTA(Tracker!$A$5:$A$200)-COUNTIF(Tracker!${col_letter}$5:${col_letter}$200,"Verified")',
        )
        style_cell(ws.cell(row=row_idx, column=4), fill_color=LIGHT_BLUE, bold=True)
        style_cell(ws.cell(row=row_idx, column=5), h_align="center")

    ws["A24"] = "Priority Mix"
    style_cell(ws["A24"], fill_color=DARK_BLUE, bold=True, font_color=WHITE, h_align="center")
    ws["B24"] = "Count"
    style_cell(ws["B24"], fill_color=DARK_BLUE, bold=True, font_color=WHITE, h_align="center")
    priorities = ["Urgent", "High", "Medium", "Low", "Done"]
    for row_idx, priority in enumerate(priorities, start=25):
        ws.cell(row=row_idx, column=1, value=priority)
        ws.cell(row=row_idx, column=2, value=f'=COUNTIF(Tracker!$F$5:$F$200,A{row_idx})')
        style_cell(ws.cell(row=row_idx, column=1), fill_color=STATUS_COLORS[priority], bold=True)
        style_cell(ws.cell(row=row_idx, column=2), h_align="center")

    status_chart = PieChart()
    status_chart.title = "Team Status Distribution"
    status_chart.height = 8
    status_chart.width = 10
    status_data = Reference(ws, min_col=2, min_row=13, max_row=20)
    status_labels = Reference(ws, min_col=1, min_row=14, max_row=20)
    status_chart.add_data(status_data, titles_from_data=True)
    status_chart.set_categories(status_labels)
    status_chart.style = 10
    status_chart.legend.position = "r"
    ws.add_chart(status_chart, "G13")

    gap_chart = BarChart()
    gap_chart.type = "bar"
    gap_chart.style = 10
    gap_chart.title = "Apps Still Needing Verification"
    gap_chart.y_axis.title = "Required Software"
    gap_chart.x_axis.title = "People"
    gap_chart.height = 8
    gap_chart.width = 10
    gap_data = Reference(ws, min_col=5, min_row=13, max_row=20)
    gap_labels = Reference(ws, min_col=4, min_row=14, max_row=20)
    gap_chart.add_data(gap_data, titles_from_data=True)
    gap_chart.set_categories(gap_labels)
    ws.add_chart(gap_chart, "G30")

    return ws


def build_guide_sheet(wb: Workbook):
    ws = wb.create_sheet("Guide")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "9BBB59"

    merge_label(ws, "A1:J1", "How To Use This Workbook", DARK_BLUE, size=16)
    merge_label(
        ws,
        "A2:J2",
        "The Tracker is the working sheet. The Dashboard is the manager view. This page explains the rules.",
        MID_BLUE,
        size=11,
        color=DARK_BLUE,
    )

    for column, width in {
        "A": 24,
        "B": 16,
        "C": 20,
        "D": 20,
        "E": 20,
        "F": 24,
        "G": 16,
        "H": 18,
        "I": 18,
        "J": 18,
    }.items():
        ws.column_dimensions[column].width = width

    ws["A4"] = "Quick Workflow"
    style_cell(ws["A4"], fill_color=DARK_BLUE, bold=True, font_color=WHITE)
    steps = [
        "1. Update the seven software columns in Tracker using the dropdowns.",
        "2. Use Workflow Demo when the salesperson shows the real process live.",
        "3. Use Manager Sign-Off only after the person is truly ready to sell.",
        "4. Sort Tracker by Priority, then Next Follow-Up, to work the list efficiently.",
        "5. Review Dashboard for readiness counts and software bottlenecks.",
    ]
    for row_idx, step in enumerate(steps, start=5):
        ws.cell(row=row_idx, column=1, value=step)
        style_cell(ws.cell(row=row_idx, column=1), fill_color=SOFT_YELLOW, wrap=True)

    ws["A12"] = "Software Status Legend"
    ws["D12"] = "Meaning"
    style_cell(ws["A12"], fill_color=DARK_BLUE, bold=True, font_color=WHITE)
    style_cell(ws["D12"], fill_color=DARK_BLUE, bold=True, font_color=WHITE)
    for row_idx, (label, _) in enumerate(SOFTWARE_STATUS_SCORES, start=13):
        ws.cell(row=row_idx, column=1, value=label)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
        ws.cell(row=row_idx, column=4, value={
            "Not Started": "No request submitted yet.",
            "Requested": "Access request sent but setup has not begun.",
            "Provisioning": "IT or the vendor is building the account.",
            "Password Reset Needed": "Account exists but the user still cannot log in.",
            "Active": "Access works but has not been verified live with the salesperson.",
            "Verified": "Salesperson logged in and confirmed the tool is usable.",
            "Training Needed": "The tool works but the salesperson needs coaching.",
            "Blocked": "A known issue is stopping progress.",
        }[label])
        style_cell(ws.cell(row=row_idx, column=1), fill_color=STATUS_COLORS[label], bold=True)
        style_cell(ws.cell(row=row_idx, column=2), fill_color=STATUS_COLORS[label], border=THIN_BORDER)
        style_cell(ws.cell(row=row_idx, column=4), wrap=True)

    ws["F12"] = "Overall Status"
    ws["H12"] = "When It Should Show Up"
    style_cell(ws["F12"], fill_color=DARK_BLUE, bold=True, font_color=WHITE)
    style_cell(ws["H12"], fill_color=DARK_BLUE, bold=True, font_color=WHITE)

    overall_definitions = [
        ("Not Started", "Nothing has moved yet."),
        ("Access In Progress", "At least one required app is still being requested or provisioned."),
        ("Access Live - Verify", "Some access is live, but not all seven apps are verified."),
        ("Needs Training", "Core access exists, but the salesperson still needs coaching."),
        ("Pending Final Review", "All seven apps are verified, but the demo or approval still needs to finish."),
        ("Ready To Sell", "Everything is verified and management approved the person."),
        ("Blocked", "A hard stop exists and this needs attention quickly."),
    ]
    for row_idx, (label, meaning) in enumerate(overall_definitions, start=13):
        ws.cell(row=row_idx, column=6, value=label)
        ws.cell(row=row_idx, column=8, value=meaning)
        style_cell(ws.cell(row=row_idx, column=6), fill_color=STATUS_COLORS[label], bold=True)
        style_cell(ws.cell(row=row_idx, column=8), wrap=True)

    ws["F24"] = "Priority"
    ws["H24"] = "Use"
    style_cell(ws["F24"], fill_color=DARK_BLUE, bold=True, font_color=WHITE)
    style_cell(ws["H24"], fill_color=DARK_BLUE, bold=True, font_color=WHITE)
    priority_definitions = [
        ("Urgent", "Work these first. They are blocked."),
        ("High", "Important follow-up is needed soon."),
        ("Medium", "Access is live, but verification still needs to happen."),
        ("Low", "Nothing urgent has started yet."),
        ("Done", "Fully ready to sell."),
    ]
    for row_idx, (label, meaning) in enumerate(priority_definitions, start=25):
        ws.cell(row=row_idx, column=6, value=label)
        ws.cell(row=row_idx, column=8, value=meaning)
        style_cell(ws.cell(row=row_idx, column=6), fill_color=STATUS_COLORS[label], bold=True)
        style_cell(ws.cell(row=row_idx, column=8), wrap=True)

    return ws


def build_workbook() -> Path:
    wb = Workbook()
    wb.properties.creator = "Codex"
    wb.properties.title = "Sales Access Control Center"
    wb.properties.subject = "Sales software readiness tracking"
    wb.properties.description = "Operational Excel workbook for tracking software access and readiness across the sales team."
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    build_lists_sheet(wb)
    build_tracker_sheet(wb)
    build_dashboard_sheet(wb)
    build_guide_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    output = build_workbook()
    print(output)

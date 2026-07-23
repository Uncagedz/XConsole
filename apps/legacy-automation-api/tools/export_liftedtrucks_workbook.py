from __future__ import annotations

import json
import re
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path
from urllib.parse import unquote, urljoin

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service


ROOT = Path(__file__).resolve().parents[1]
REPORTS_DIR = ROOT / "runtime" / "reports"
DETAIL_CACHE_DIR = REPORTS_DIR / "liftedtrucks_detail_cache"
ANALYSIS_GLOB = "liftedtrucks_inventory_analysis_*.json"
INVENTORY_URLS = [
    "https://www.liftedtrucks.com/used-inventory/ram-trucks.htm",
    "https://www.liftedtrucks.com/used-inventory/jeep.htm",
]
CHROMEDRIVER = ROOT / "automation" / "facebook-marketplace-lister" / "drivers" / "chromedriver.exe"
CHROME_CANDIDATES = [
    ROOT / "automation" / "facebook-marketplace-lister" / "chrome-for-testing" / "chrome-win64" / "chrome.exe",
    Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
    Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
]
TODAY = date(2026, 4, 1)


def _latest_analysis_report() -> Path:
    matches = sorted(REPORTS_DIR.glob(ANALYSIS_GLOB), key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError(f"No analysis report found in {REPORTS_DIR}")
    return matches[0]


def _make_browser() -> webdriver.Chrome:
    opts = webdriver.ChromeOptions()
    opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--window-size=1440,1400")
    for candidate in CHROME_CANDIDATES:
        if candidate.exists():
            opts.binary_location = str(candidate)
            break
    return webdriver.Chrome(service=Service(str(CHROMEDRIVER.resolve())), options=opts)


def _page_payload(source: str) -> dict:
    match = re.search(
        r'fetch\("/api/widget/ws-inv-data/getInventory".*?decodeURI\("([^"]+)"\)',
        source,
        re.S,
    )
    if not match:
        raise RuntimeError("Could not find inventory API payload in page source")
    return json.loads(unquote(match.group(1)))


def _browser_fetch_inventory(driver: webdriver.Chrome, payload: dict) -> dict:
    script = """
    return fetch("/api/widget/ws-inv-data/getInventory", {
      method: "POST",
      headers: {"Content-Type": "application/json"},
      body: JSON.stringify(arguments[0]),
    }).then(r => r.text().then(t => ({status: r.status, text: t})))
      .catch(e => ({status: 0, text: String(e)}));
    """
    result = driver.execute_script(script, payload)
    if int(result.get("status") or 0) != 200:
        raise RuntimeError(f"Inventory fetch failed: {result}")
    return json.loads(result["text"])


def _absolute_listing_url(item: dict) -> str:
    return urljoin("https://www.liftedtrucks.com", str(item.get("link") or ""))


def _parse_inventory_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value.strip(), "%m/%d/%Y").date()
    except ValueError:
        return None


def _days_in_stock(inventory_date: date | None) -> int | None:
    if not inventory_date:
        return None
    return max((TODAY - inventory_date).days, 0)


def _slug_for_listing_url(listing_url: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", listing_url.rsplit("/", 1)[-1])


def _load_detail_page_text(listing_url: str) -> str:
    cache_path = DETAIL_CACHE_DIR / f"{_slug_for_listing_url(listing_url)}.txt"
    if not cache_path.exists():
        return ""
    return cache_path.read_text(encoding="utf-8")


def _extract_views_7d(detail_text: str) -> int | None:
    match = re.search(r"(\d+)\s+views?\s+in\s+the\s+past\s+7\s+days", detail_text, re.I)
    return int(match.group(1)) if match else None


def _extract_page_urls(driver: webdriver.Chrome, inventory_url: str) -> list[str]:
    driver.get(inventory_url)
    import time
    time.sleep(6)
    starts = sorted(
        {
            int(match.group(1))
            for anchor in driver.find_elements("tag name", "a")
            for href in [anchor.get_attribute("href") or ""]
            for match in [re.search(r"[?&]start=(\d+)", href)]
            if match
        }
    )
    base = inventory_url.split("?", 1)[0]
    if not starts:
        return [base]
    step = min(start for start in starts if start > 0) if any(start > 0 for start in starts) else 27
    last = max(starts)
    return [base] + [f"{base}?start={start}" for start in range(step, last + 1, step)]


def collect_inventory_metrics() -> dict[str, dict]:
    metrics: dict[str, dict] = {}
    driver = _make_browser()
    import time
    try:
        for inventory_url in INVENTORY_URLS:
            for page_url in _extract_page_urls(driver, inventory_url):
                driver.get(page_url)
                time.sleep(5)
                payload = _page_payload(driver.page_source)
                data = _browser_fetch_inventory(driver, payload)
                for item in data.get("inventory") or []:
                    listing_url = _absolute_listing_url(item)
                    inventory_date = _parse_inventory_date(item.get("inventoryDate"))
                    detail_text = _load_detail_page_text(listing_url)
                    views_7d = _extract_views_7d(detail_text)
                    metrics[listing_url] = {
                        "inventory_page_url": inventory_url,
                        "inventory_date": inventory_date.isoformat() if inventory_date else None,
                        "days_in_stock": _days_in_stock(inventory_date),
                        "website_views_7d": views_7d,
                        "website_clicks_public": views_7d,
                        "public_engagement_velocity_7d": round(views_7d / 7.0, 2) if views_7d is not None else None,
                        "public_views_per_live_day": (
                            round(views_7d / max(1, min(_days_in_stock(inventory_date) or 1, 7)), 2)
                            if views_7d is not None
                            else None
                        ),
                        "inventory_api_title": item.get("title"),
                        "inventory_api_stock_number": item.get("stockNumber"),
                        "inventory_api_vehicle_price": (
                            item.get("pricing", {}).get("salePrice")
                            or item.get("pricing", {}).get("internetPrice")
                            or item.get("trackingPricing", {}).get("internetPrice")
                        ),
                        "inventory_api_featured_price": item.get("trackingPricing", {}).get("askingPrice"),
                    }
    finally:
        driver.quit()
    return metrics


def _flatten_vehicle(vehicle: dict, metrics: dict[str, dict]) -> dict:
    raw = deepcopy(vehicle["section_1_raw_vehicle_data"])
    mods = vehicle["section_2_modification_detection"]
    reverse = vehicle["section_3_build_reverse_engineering"]
    perf = vehicle["section_4_performance_and_use_analysis"]
    repl = vehicle["section_5_replication_guide"]
    notes = vehicle["section_6_data_extraction_notes"]
    listing_url = raw["listing_url"]
    metric = metrics.get(listing_url, {})

    lift = mods["lift_kit"]
    wheel = mods["wheels"]
    tire = mods["tires"]
    susp = mods["suspension_components"]

    row = {
        "vehicle_key": vehicle["vehicle_key"],
        "year": raw.get("year"),
        "make": raw.get("make"),
        "model": raw.get("model"),
        "trim": raw.get("trim"),
        "vin": raw.get("vin"),
        "stock_number": raw.get("stock_number"),
        "body_style": raw.get("body_style"),
        "condition": raw.get("condition"),
        "mileage": raw.get("mileage"),
        "featured_price_usd": raw.get("featured_price_usd"),
        "vehicle_price_usd": raw.get("vehicle_price_usd"),
        "aftermarket_accessories_usd": raw.get("aftermarket_price_value_usd"),
        "engine": raw.get("engine"),
        "transmission": raw.get("transmission"),
        "drivetrain": raw.get("drivetrain"),
        "exterior_color": raw.get("exterior_color"),
        "interior_color": raw.get("interior_color"),
        "listing_url": listing_url,
        "inventory_source_url": raw.get("inventory_source_url"),
        "carfax_url": raw.get("carfax_url"),
        "image_count": raw.get("image_count"),
        "inventory_date": metric.get("inventory_date"),
        "days_in_stock": metric.get("days_in_stock"),
        "website_views_7d": metric.get("website_views_7d"),
        "website_clicks_public": metric.get("website_clicks_public"),
        "public_engagement_velocity_7d": metric.get("public_engagement_velocity_7d"),
        "public_views_per_live_day": metric.get("public_views_per_live_day"),
        "lift_type": lift.get("type"),
        "lift_height_inches": lift.get("estimated_height_inches"),
        "lift_brand": lift.get("likely_brand"),
        "lift_brand_model": lift.get("brand_model"),
        "lift_evidence": lift.get("evidence"),
        "shock_package": ", ".join(susp.get("shocks_struts") or []),
        "wheel_size_inches": wheel.get("size_inches"),
        "wheel_style": wheel.get("style"),
        "wheel_offset_estimate": wheel.get("offset_estimate"),
        "tire_size": tire.get("size"),
        "tire_type": tire.get("type"),
        "tire_brand": tire.get("brand"),
        "tire_brand_model": tire.get("brand_model"),
        "other_mods": " | ".join(mods.get("other_mods") or []),
        "most_likely_lift_setup": reverse.get("most_likely_lift_setup"),
        "required_supporting_mods": " | ".join(reverse.get("required_supporting_mods") or []),
        "likely_corners_cut": " | ".join(reverse.get("likely_corners_cut") or []),
        "build_quality_tier": reverse.get("build_quality_tier"),
        "street_vs_offroad_balance": perf.get("street_vs_offroad_balance"),
        "ride_quality_estimate": perf.get("ride_quality_estimate"),
        "weak_points": " | ".join(perf.get("weak_points") or []),
        "long_term_reliability_concerns": " | ".join(perf.get("long_term_reliability_concerns") or []),
        "parts_list_prioritized": " | ".join(repl.get("parts_list_prioritized") or []),
        "install_difficulty_1_to_10": repl.get("install_difficulty_1_to_10"),
        "what_to_upgrade_or_change": " | ".join(repl.get("what_to_upgrade_or_change") or []),
        "confidence_level": notes.get("confidence_level"),
        "missing_data": " | ".join(notes.get("missing_data") or []),
        "dealer_notes_detailed_paraphrase": raw.get("dealer_notes_detailed_paraphrase"),
        "public_engagement_note": (
            "Website only exposes public 'views in the past 7 days'. No public save / lead counts were exposed."
        ),
    }
    return row


def _rank_rows(rows: list[dict], metric_name: str, rank_name: str) -> None:
    ranked = sorted(
        [row for row in rows if row.get(metric_name) is not None],
        key=lambda row: (row.get(metric_name) or 0, row.get("featured_price_usd") or 0),
        reverse=True,
    )
    for index, row in enumerate(ranked, start=1):
        row[rank_name] = index
    for row in rows:
        row.setdefault(rank_name, None)


def _autosize_sheet(ws) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:150]]
        max_len = max((len(value) for value in values), default=0)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 48)


def _write_sheet(ws, rows: list[dict]) -> None:
    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, len(rows) + 1)}"
    for row in rows:
        ws.append([row.get(header) for header in headers])
    _autosize_sheet(ws)


def export_workbook(rows: list[dict], output_xlsx: Path, output_csv: Path) -> None:
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All Vehicles"
    _write_sheet(ws_all, rows)

    top_clicks = sorted(
        rows,
        key=lambda row: (
            row.get("website_clicks_public") if row.get("website_clicks_public") is not None else -1,
            row.get("featured_price_usd") if row.get("featured_price_usd") is not None else -1,
        ),
        reverse=True,
    )
    ws_clicks = wb.create_sheet("Top Clicks")
    _write_sheet(ws_clicks, top_clicks)

    top_engagement = sorted(
        rows,
        key=lambda row: (
            row.get("public_views_per_live_day") if row.get("public_views_per_live_day") is not None else -1,
            row.get("website_clicks_public") if row.get("website_clicks_public") is not None else -1,
        ),
        reverse=True,
    )
    ws_engagement = wb.create_sheet("Top Engagement")
    _write_sheet(ws_engagement, top_engagement)

    ws_missing = wb.create_sheet("Missing Metrics")
    missing_rows = [row for row in rows if row.get("website_clicks_public") is None or row.get("days_in_stock") is None]
    _write_sheet(ws_missing, missing_rows if missing_rows else rows[:1])

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    import csv

    headers = list(rows[0].keys()) if rows else []
    with output_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    analysis_path = _latest_analysis_report()
    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    metrics = collect_inventory_metrics()
    rows = [_flatten_vehicle(vehicle, metrics) for vehicle in analysis["vehicles"]]
    _rank_rows(rows, "website_clicks_public", "click_rank_7d")
    _rank_rows(rows, "public_views_per_live_day", "engagement_rank_velocity")

    datestamp = TODAY.strftime("%Y%m%d")
    output_xlsx = Path(rf"D:\liftedtrucks_inventory_intelligence_{datestamp}.xlsx")
    output_csv = Path(rf"D:\liftedtrucks_inventory_intelligence_{datestamp}.csv")
    export_workbook(rows, output_xlsx, output_csv)

    summary = {
        "ok": True,
        "analysis_source": str(analysis_path),
        "xlsx": str(output_xlsx),
        "csv": str(output_csv),
        "row_count": len(rows),
        "with_days_in_stock": sum(1 for row in rows if row.get("days_in_stock") is not None),
        "with_public_views_7d": sum(1 for row in rows if row.get("website_clicks_public") is not None),
    }
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

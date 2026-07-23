from __future__ import annotations

import csv
import json
import re
from datetime import date
from pathlib import Path
from statistics import mean

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
REPORTS_DIR = ROOT / "runtime" / "reports"
DETAIL_CACHE_DIR = REPORTS_DIR / "liftedtrucks_detail_cache"
TODAY = date(2026, 4, 1)

PREMIUM_PART_BRANDS = [
    "bds",
    "carli",
    "fox",
    "icon",
    "king",
    "method",
    "fuel",
    "nitto",
    "toyo",
    "readylift",
    "teraflex",
    "metalcloak",
    "black rhino",
    "kmc",
    "dv8",
    "rival",
    "cognito",
]


def _latest_intelligence_csv() -> Path:
    candidates = sorted(Path("D:/").glob("liftedtrucks_inventory_intelligence_*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
    if candidates:
        return candidates[0]
    raise FileNotFoundError("Could not find liftedtrucks_inventory_intelligence_*.csv on D:\\")


def _slug_for_listing_url(listing_url: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", listing_url.rsplit("/", 1)[-1])


def _load_detail_text(listing_url: str) -> str:
    path = DETAIL_CACHE_DIR / f"{_slug_for_listing_url(listing_url)}.txt"
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore")


def _to_float(value: str | None) -> float | None:
    if value in (None, "", "None"):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def _to_int(value: str | None) -> int | None:
    number = _to_float(value)
    return int(number) if number is not None else None


def _is_unknown(value: str | None) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    if not text:
        return True
    lowered = text.lower()
    bad_fragments = [
        "unknown",
        "not explicitly",
        "no explicit",
        "none clearly stated",
        "[inferred] none",
    ]
    return any(fragment in lowered for fragment in bad_fragments)


def _clean_label(value: str | None) -> str:
    if _is_unknown(value):
        return ""
    text = re.sub(r"\[INFERRED\]\s*", "", str(value)).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _extract_carfax_signals(detail_text: str, dealer_notes: str) -> dict:
    text = f"{detail_text}\n{dealer_notes or ''}"
    badge_match = re.search(r"(?:Show Me the )?(CARFAX[^\\\]\n]*?Badge)", text, re.I)
    badge_text = badge_match.group(1).strip() if badge_match else ""
    badge_text = badge_text.replace("Show Me the ", "")

    clean_carfax = bool(re.search(r"\bclean carfax\b", text, re.I))
    no_accidents = bool(re.search(r"\bno accidents reported\b", text, re.I))
    one_owner = bool(re.search(r"\bone owner\b", text, re.I)) or bool(re.search(r"1-Owner", badge_text, re.I))
    good_value = bool(re.search(r"\bgood value\b", badge_text, re.I))
    fair_value = bool(re.search(r"\bfair value\b", badge_text, re.I))
    carfax_present = bool(badge_text)

    summary_parts: list[str] = []
    if one_owner:
        summary_parts.append("1-owner")
    if clean_carfax:
        summary_parts.append("clean Carfax")
    if no_accidents:
        summary_parts.append("no accidents reported")
    if good_value:
        summary_parts.append("good value badge")
    elif fair_value:
        summary_parts.append("fair value badge")
    elif carfax_present and not summary_parts:
        summary_parts.append("Carfax badge present")
    elif carfax_present:
        summary_parts.append("Carfax badge present")
    if not summary_parts:
        summary_parts.append("No Carfax detail extracted")

    score = 0
    if carfax_present:
        score += 1
    if one_owner:
        score += 4
    if clean_carfax:
        score += 8
    if no_accidents:
        score += 6
    if good_value:
        score += 2
    if fair_value:
        score -= 2
    score = max(0, min(score, 20))

    return {
        "badge_text": badge_text or "",
        "summary": "; ".join(summary_parts),
        "score": score,
        "one_owner": one_owner,
        "clean_carfax": clean_carfax,
        "no_accidents": no_accidents,
        "good_value": good_value,
        "fair_value": fair_value,
        "present": carfax_present,
        "strong_signal": clean_carfax or no_accidents or one_owner,
    }


def _parts_summary(row: dict) -> str:
    parts: list[str] = []

    lift = _clean_label(row.get("lift_brand_model"))
    if lift:
        height = _clean_label(row.get("lift_height_inches"))
        if height and not re.search(r"\d", lift):
            parts.append(f"{height} lift ({lift})")
        else:
            parts.append(lift)

    shocks = _clean_label(row.get("shock_package"))
    if shocks and shocks.lower() not in " ".join(parts).lower():
        parts.append(shocks)

    wheel_size = _clean_label(row.get("wheel_size_inches"))
    wheel_style = _clean_label(row.get("wheel_style"))
    if wheel_size or wheel_style:
        wheel_text = " ".join(part for part in [f'{wheel_size}"' if wheel_size else "", wheel_style] if part).strip()
        if wheel_text:
            parts.append(wheel_text)

    tire_size = _clean_label(row.get("tire_size"))
    tire_model = _clean_label(row.get("tire_brand_model"))
    tire_text = " ".join(part for part in [tire_size, tire_model] if part).strip()
    if tire_text:
        parts.append(tire_text)

    other_mods = [item.strip() for item in str(row.get("other_mods") or "").split("|") if item.strip()]
    for item in other_mods:
        clean_item = _clean_label(item)
        if clean_item and clean_item.lower() not in " ".join(parts).lower():
            parts.append(clean_item)

    if not parts:
        raw_parts = [item.strip() for item in str(row.get("parts_list_prioritized") or "").split("|") if item.strip()]
        parts = [_clean_label(item) for item in raw_parts if _clean_label(item)]

    return "; ".join(parts[:6]) if parts else "Not clearly disclosed"


def _click_percentile(rank: int | None, population: int) -> float:
    if rank is None or population <= 1:
        return 0.15
    return max(0.0, 1.0 - ((rank - 1) / (population - 1)))


def _year_score(year: int | None, min_year: int, max_year: int) -> float:
    if year is None or min_year == max_year:
        return 7.5
    return round(15 * ((year - min_year) / (max_year - min_year)), 1)


def _engagement_score(click_rank: int | None, engagement_rank: int | None, click_population: int, engagement_population: int) -> float:
    click_pct = _click_percentile(click_rank, click_population)
    engagement_pct = _click_percentile(engagement_rank, engagement_population)
    return round((12.5 * click_pct) + (12.5 * engagement_pct), 1)


def _supply_score(days_in_stock: int | None) -> float:
    if days_in_stock is None:
        return 8.0
    if days_in_stock <= 14:
        return 20.0
    if days_in_stock <= 30:
        return 17.0
    if days_in_stock <= 45:
        return 14.0
    if days_in_stock <= 60:
        return 11.0
    if days_in_stock <= 90:
        return 8.0
    if days_in_stock <= 120:
        return 5.0
    return 2.0


def _build_score(build_quality_tier: str | None, parts_summary: str) -> float:
    lowered = str(build_quality_tier or "").lower()
    if "upper-mid to premium" in lowered:
        score = 8.0
    elif "mid-tier" in lowered:
        score = 6.0
    else:
        score = 5.0

    parts_lower = parts_summary.lower()
    premium_hits = sum(1 for brand in PREMIUM_PART_BRANDS if brand in parts_lower)
    if premium_hits >= 3:
        score += 2.0
    elif premium_hits >= 1:
        score += 1.0
    return min(score, 10.0)


def _aftermarket_score(aftermarket_price: float | None, vehicle_price: float | None, parts_summary: str) -> float:
    if aftermarket_price and vehicle_price and vehicle_price > 0:
        ratio = aftermarket_price / vehicle_price
        if ratio >= 0.20:
            return 10.0
        if ratio >= 0.15:
            return 8.0
        if ratio >= 0.10:
            return 6.0
        if ratio >= 0.05:
            return 4.0
        return 2.0

    if parts_summary != "Not clearly disclosed":
        parts_lower = parts_summary.lower()
        if any(brand in parts_lower for brand in PREMIUM_PART_BRANDS):
            return 4.0
        return 2.0
    return 1.0


def _confidence_penalty(confidence: str | None) -> float:
    lowered = str(confidence or "").lower()
    if lowered == "low":
        return 4.0
    if lowered == "medium":
        return 2.0
    return 0.0


def _rank_bucket(rank: int | None) -> str:
    if rank is None:
        return "No public metric"
    if rank <= 10:
        return "Top 10"
    if rank <= 25:
        return "Top 25"
    if rank <= 50:
        return "Top 50"
    return ""


def _suggestion(score: float, days_in_stock: int | None, carfax: dict, click_rank: int | None, engagement_rank: int | None) -> str:
    if score >= 80:
        return "Yes - top candidate"
    if score >= 70:
        return "Yes - strong watchlist"
    if score >= 62 and days_in_stock is not None and days_in_stock > 60 and (click_rank or 999) <= 50:
        return "Maybe - aged but market is still looking"
    if score >= 62 and carfax["strong_signal"]:
        return "Maybe - needs manual review"
    return "Lower priority"


def _reason(row: dict, carfax: dict, score: float) -> str:
    positives: list[str] = []
    watch_items: list[str] = []

    year = _to_int(row.get("year"))
    days_in_stock = _to_int(row.get("days_in_stock"))
    click_rank = _to_int(row.get("click_rank_7d"))
    engagement_rank = _to_int(row.get("engagement_rank_velocity"))
    aftermarket_price = _to_float(row.get("aftermarket_accessories_usd"))
    vehicle_price = _to_float(row.get("vehicle_price_usd")) or _to_float(row.get("featured_price_usd"))
    confidence = str(row.get("confidence_level") or "")

    if year is not None and year >= 2024:
        positives.append("newer model year")
    if click_rank is not None and click_rank <= 25:
        positives.append("strong click traffic")
    if engagement_rank is not None and engagement_rank <= 25:
        positives.append("strong engagement velocity")
    if carfax["clean_carfax"]:
        positives.append("clean Carfax wording")
    elif carfax["one_owner"]:
        positives.append("1-owner Carfax signal")
    if days_in_stock is not None and days_in_stock <= 30:
        positives.append("fresh inventory")
    if aftermarket_price and vehicle_price and vehicle_price > 0 and (aftermarket_price / vehicle_price) >= 0.10:
        positives.append("meaningful aftermarket value")

    if days_in_stock is not None and days_in_stock > 90:
        watch_items.append("aged stock")
    if not carfax["strong_signal"]:
        watch_items.append("limited Carfax detail")
    if click_rank is not None and click_rank > 100 and (engagement_rank or 999) > 100:
        watch_items.append("light recent traffic")
    if confidence.lower() == "low":
        watch_items.append("build details less certain")
    if aftermarket_price is None:
        watch_items.append("aftermarket price not disclosed")

    positive_text = ", ".join(positives[:3]) if positives else "mixed signals"
    watch_text = ", ".join(watch_items[:2]) if watch_items else "no major warning flags"
    return f"Why: {positive_text}. Watch: {watch_text}. Score {score:.1f}/100."


def _autosize_sheet(ws) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:200]]
        width = min(max((max((len(value) for value in values), default=0) + 2), 12), 44)
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_sheet(ws, rows: list[dict]) -> None:
    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, len(rows) + 1)}"
    for row in rows:
        ws.append([row.get(header) for header in headers])
    _autosize_sheet(ws)


def _score_color_rules(ws, header_map: dict[str, int], row_count: int) -> None:
    for column_name in ["overall_score", "score_carfax", "score_engagement", "score_supply"]:
        if column_name in header_map:
            col = get_column_letter(header_map[column_name])
            ws.conditional_formatting.add(
                f"{col}2:{col}{row_count + 1}",
                ColorScaleRule(start_type="num", start_value=0, start_color="F87171", mid_type="num", mid_value=50, mid_color="FCD34D", end_type="num", end_value=100, end_color="34D399"),
            )
    if "days_in_stock" in header_map:
        col = get_column_letter(header_map["days_in_stock"])
        ws.conditional_formatting.add(
            f"{col}2:{col}{row_count + 1}",
            ColorScaleRule(start_type="num", start_value=0, start_color="34D399", mid_type="num", mid_value=45, mid_color="FCD34D", end_type="num", end_value=180, end_color="F87171"),
        )


def _summary_rows(rows: list[dict]) -> list[tuple[str, str]]:
    scores = [row["overall_score"] for row in rows]
    click_rows = [row for row in rows if row["website_clicks_7d"] is not None]
    engagement_rows = [row for row in rows if row["engagement_per_live_day"] is not None]
    top_score = rows[0]
    top_click = max(click_rows, key=lambda row: row["website_clicks_7d"]) if click_rows else None
    top_engagement = max(engagement_rows, key=lambda row: row["engagement_per_live_day"]) if engagement_rows else None

    return [
        ("Generated On", TODAY.isoformat()),
        ("Total Vehicles", str(len(rows))),
        ("Average Score", f"{mean(scores):.1f}"),
        ("Average Days In Stock", f"{mean([row['days_in_stock'] for row in rows if row['days_in_stock'] is not None]):.1f}"),
        ("Vehicles With Carfax Signal", str(sum(1 for row in rows if row["carfax_summary"] != "No Carfax detail extracted"))),
        ("Vehicles With Public Click Data", str(len(click_rows))),
        ("Top Score VIN", f"{top_score['vin']} | {top_score['year']} {top_score['make']} {top_score['model']} {top_score['trim']} | {top_score['overall_score']:.1f}"),
        ("Top Click VIN", f"{top_click['vin']} | {top_click['website_clicks_7d']} clicks" if top_click else ""),
        ("Top Engagement VIN", f"{top_engagement['vin']} | {top_engagement['engagement_per_live_day']:.2f}/day" if top_engagement else ""),
        ("Carfax Note", "Summary uses dealer-page Carfax badges and dealer wording. Full Carfax reports were not directly scraped."),
        ("Days Of Supply Note", "Used as a site-age proxy from inventory date because true market turn-rate is not exposed publicly."),
    ]


def _scoring_model_rows() -> list[dict]:
    return [
        {"component": "score_year", "max_points": 15, "logic": "Newer model years score higher across the observed inventory range."},
        {"component": "score_carfax", "max_points": 20, "logic": "Dealer-page Carfax badge plus wording such as 1-owner, clean Carfax, no accidents reported, good/fair value."},
        {"component": "score_engagement", "max_points": 25, "logic": "Blend of click rank and engagement-velocity rank from public 7-day website views."},
        {"component": "score_supply", "max_points": 20, "logic": "Favors fresher inventory; older site age reduces points."},
        {"component": "score_build", "max_points": 10, "logic": "Uses build tier plus premium aftermarket brand detection."},
        {"component": "score_aftermarket", "max_points": 10, "logic": "Uses disclosed aftermarket value relative to base vehicle price, with a small fallback for clearly disclosed premium parts."},
        {"component": "score_confidence_penalty", "max_points": "-4", "logic": "Subtracts points when build extraction confidence is medium or low."},
        {"component": "overall_score", "max_points": 100, "logic": "Sum of positive components minus data-confidence penalty, capped at 0-100."},
    ]


def build_dashboard_rows() -> list[dict]:
    source_csv = _latest_intelligence_csv()
    with source_csv.open("r", newline="", encoding="utf-8") as handle:
        source_rows = list(csv.DictReader(handle))

    years = [_to_int(row.get("year")) for row in source_rows if _to_int(row.get("year")) is not None]
    min_year = min(years)
    max_year = max(years)
    click_population = sum(1 for row in source_rows if _to_int(row.get("click_rank_7d")) is not None)
    engagement_population = sum(1 for row in source_rows if _to_int(row.get("engagement_rank_velocity")) is not None)

    dashboard_rows: list[dict] = []
    for row in source_rows:
        detail_text = _load_detail_text(str(row.get("listing_url") or ""))
        carfax = _extract_carfax_signals(detail_text, str(row.get("dealer_notes_detailed_paraphrase") or ""))
        parts_summary = _parts_summary(row)

        year = _to_int(row.get("year"))
        days_in_stock = _to_int(row.get("days_in_stock"))
        vehicle_price = _to_float(row.get("vehicle_price_usd")) or _to_float(row.get("featured_price_usd"))
        aftermarket_price = _to_float(row.get("aftermarket_accessories_usd"))
        total_price = _to_float(row.get("featured_price_usd"))
        click_rank = _to_int(row.get("click_rank_7d"))
        engagement_rank = _to_int(row.get("engagement_rank_velocity"))
        clicks_7d = _to_int(row.get("website_clicks_public"))
        engagement_per_day = _to_float(row.get("public_views_per_live_day"))

        score_year = _year_score(year, min_year, max_year)
        score_carfax = float(carfax["score"])
        score_engagement = _engagement_score(click_rank, engagement_rank, click_population, engagement_population)
        score_supply = _supply_score(days_in_stock)
        score_build = _build_score(row.get("build_quality_tier"), parts_summary)
        score_aftermarket = _aftermarket_score(aftermarket_price, vehicle_price, parts_summary)
        score_penalty = _confidence_penalty(row.get("confidence_level"))
        overall_score = round(max(0.0, min(100.0, score_year + score_carfax + score_engagement + score_supply + score_build + score_aftermarket - score_penalty)), 1)

        dashboard_row = {
            "overall_score": overall_score,
            "suggested_for_me": _suggestion(overall_score, days_in_stock, carfax, click_rank, engagement_rank),
            "score_reason": _reason(row, carfax, overall_score),
            "year": year,
            "make": row.get("make"),
            "model": row.get("model"),
            "trim": row.get("trim"),
            "vin": row.get("vin"),
            "vehicle_price_usd": vehicle_price,
            "aftermarket_price_usd": aftermarket_price,
            "asking_total_usd": total_price,
            "aftermarket_parts": parts_summary,
            "carfax_summary": carfax["summary"],
            "carfax_badge": carfax["badge_text"],
            "days_in_stock": days_in_stock,
            "days_of_supply_proxy": days_in_stock,
            "website_clicks_7d": clicks_7d,
            "engagement_per_live_day": engagement_per_day,
            "top_clicks": _rank_bucket(click_rank),
            "top_engagement": _rank_bucket(engagement_rank),
            "click_rank_7d": click_rank,
            "engagement_rank": engagement_rank,
            "build_quality_tier": _clean_label(row.get("build_quality_tier")) or "Unknown",
            "confidence_level": row.get("confidence_level"),
            "score_year": score_year,
            "score_carfax": score_carfax,
            "score_engagement": score_engagement,
            "score_supply": score_supply,
            "score_build": score_build,
            "score_aftermarket": score_aftermarket,
            "score_confidence_penalty": score_penalty,
        }
        dashboard_rows.append(dashboard_row)

    dashboard_rows.sort(key=lambda item: (item["overall_score"], item["website_clicks_7d"] or -1, item["year"] or 0), reverse=True)
    return dashboard_rows


def export_dashboard(rows: list[dict], output_xlsx: Path, output_csv: Path) -> None:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Value"])
    for metric, value in _summary_rows(rows):
        ws_summary.append([metric, value])
    for cell in ws_summary[1]:
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.font = Font(color="FFFFFF", bold=True)
    ws_summary.freeze_panes = "A2"
    _autosize_sheet(ws_summary)

    ws_dashboard = wb.create_sheet("Dashboard")
    _write_sheet(ws_dashboard, rows)
    header_map = {cell.value: idx for idx, cell in enumerate(ws_dashboard[1], start=1)}
    _score_color_rules(ws_dashboard, header_map, len(rows))

    ws_top = wb.create_sheet("Top Picks")
    _write_sheet(ws_top, rows[:50])

    aged_rows = [row for row in rows if row.get("days_in_stock") is not None and row["days_in_stock"] >= 60]
    aged_rows.sort(key=lambda item: (item["website_clicks_7d"] or -1, item["overall_score"]), reverse=True)
    ws_aged = wb.create_sheet("Aged Units")
    _write_sheet(ws_aged, aged_rows if aged_rows else rows[:1])

    ws_model = wb.create_sheet("Scoring Model")
    _write_sheet(ws_model, _scoring_model_rows())

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    headers = list(rows[0].keys()) if rows else []
    with output_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    rows = build_dashboard_rows()
    datestamp = TODAY.strftime("%Y%m%d")
    output_xlsx = Path(rf"D:\liftedtrucks_inventory_dashboard_{datestamp}.xlsx")
    output_csv = Path(rf"D:\liftedtrucks_inventory_dashboard_{datestamp}.csv")
    export_dashboard(rows, output_xlsx, output_csv)

    summary = {
        "ok": True,
        "xlsx": str(output_xlsx),
        "csv": str(output_csv),
        "row_count": len(rows),
        "top_vin": rows[0]["vin"] if rows else None,
        "top_score": rows[0]["overall_score"] if rows else None,
    }
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
